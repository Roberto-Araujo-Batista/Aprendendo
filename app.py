import openpyxl

#criando arquivo excel
planilha = openpyxl.Workbook()

#como visualizar pastas existentes
print(planilha.sheetnames)

#como criar uma página
planilha.create_sheet('Enviar')

#como selecionar uma página
enviar = planilha['Enviar']

enviar.append(['NOME', 'IDADE', 'CARGO'])
enviar.append(['roberto', '22', 'suporte ti'])
enviar.append(['gabriel', '5', 'estudante'])
enviar.append(['william', '9', 'estudante'])



#salvar planilha
planilha.save('dados.xlsx')



